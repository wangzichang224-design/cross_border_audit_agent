# -*- coding: utf-8 -*-
"""Tests for the neutral C cash workpaper benchmark and filler."""

from __future__ import annotations

import os
import shutil
import sys
import unittest
import uuid
from contextlib import contextmanager
from datetime import date
from pathlib import Path
from zipfile import ZipFile


REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from openpyxl import load_workbook

from benchmarks.agent import cell_map as cm
from benchmarks.agent import cell_map_clean as clean
from benchmarks.agent.cash_workpaper_filler import (
    DEFAULT_CUTOFF_WINDOW,
    DEFAULT_RISK_LEVELS,
    FIXED_GL_ACCOUNTS,
    build_fill_context,
    fill_cash_workpaper,
    write_cells_to_workbook,
)
from benchmarks.agent.materials_loader import load_case_materials
from benchmarks.template_builder.build_optimized_cash_template import build_template


CASE_001_DIR = REPO_ROOT / "benchmarks" / "materials" / "case_001_minimal"
OPTIMIZED_TEMPLATE = (
    REPO_ROOT
    / "outputs"
    / "clean_templates"
    / "C_货币资金审计底稿_核心优化版_CN_CAS.xlsx"
)


@contextmanager
def writable_temp_dir():
    """Create a writable temp directory under the repo for sandboxed Windows runs."""
    root = REPO_ROOT / ".test_tmp"
    root.mkdir(exist_ok=True)
    path = root / uuid.uuid4().hex
    path.mkdir()
    try:
        yield path
    finally:
        shutil.rmtree(path, ignore_errors=True)


def _find_template() -> Path:
    override = os.environ.get("C_CASH_TEMPLATE")
    if override:
        path = Path(override)
        if path.exists() and path.is_file():
            return path
        raise unittest.SkipTest(f"C_CASH_TEMPLATE does not exist: {path}")

    if not OPTIMIZED_TEMPLATE.exists():
        build_template()
    return OPTIMIZED_TEMPLATE


def _formula_error_hits(workbook):
    errors = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A")
    hits = []
    for ws in workbook.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and any(error in value for error in errors):
                    hits.append((ws.title, cell.coordinate, value))
    return hits


class TestMaterialsLoader(unittest.TestCase):
    def setUp(self):
        self.materials = load_case_materials(CASE_001_DIR)

    def test_metadata(self):
        meta = self.materials.meta
        self.assertEqual(meta.case_id, "case_001_minimal")
        self.assertEqual(meta.client_name, "星辰跨境科技(深圳)有限公司")
        self.assertEqual(meta.period_end, date(2025, 12, 31))
        self.assertEqual(meta.te, 5_000_000.0)
        self.assertEqual(meta.sad, 250_000.0)
        self.assertEqual(meta.gaap, "企业会计准则")
        self.assertEqual(meta.currency, "CNY")

    def test_period_summary(self):
        rows = self.materials.period_summary
        self.assertEqual(len(rows), 4)
        bank_total = sum(
            row.period_end_balance_local
            for row in rows
            if row.account_name == "银行存款"
        )
        self.assertAlmostEqual(bank_total, 7_545_565.0)

    def test_confirmations(self):
        restricted = [
            row for row in self.materials.confirmations if row.restricted_amount > 0
        ]
        self.assertEqual(len(restricted), 2)
        self.assertAlmostEqual(
            sum(row.restricted_amount for row in restricted),
            1_000_000.0,
        )

    def test_reconciliation_items(self):
        items = self.materials.reconciliation_items
        self.assertEqual(len(items), 2)
        self.assertEqual({item.category for item in items}, {"book_plus", "bank_minus"})


class TestFillContextBuilder(unittest.TestCase):
    def setUp(self):
        self.materials = load_case_materials(CASE_001_DIR)
        self.ctx = build_fill_context(self.materials)

    def test_header_fields_propagate(self):
        self.assertEqual(self.ctx.company_name, "星辰跨境科技(深圳)有限公司")
        self.assertEqual(self.ctx.period_end, date(2025, 12, 31))
        self.assertEqual(self.ctx.te, 5_000_000.0)
        self.assertEqual(self.ctx.sad, 250_000.0)
        self.assertEqual(self.ctx.currency, "CNY")

    def test_risk_levels_use_defaults(self):
        self.assertEqual(self.ctx.risk_levels, DEFAULT_RISK_LEVELS)
        for level in self.ctx.risk_levels.values():
            self.assertIn(level, cm.ALLOWED_RISK_LEVELS)

    def test_gl_rows_aggregate_by_account(self):
        gl = self.ctx.lead_gl_rows
        self.assertEqual(set(gl.keys()), set(FIXED_GL_ACCOUNTS))
        self.assertAlmostEqual(gl["库存现金"].book_value_unaudited, 32_500.0)
        self.assertAlmostEqual(gl["银行存款"].book_value_unaudited, 7_545_565.0)
        self.assertAlmostEqual(gl["其他货币资金"].book_value_unaudited, 800_000.0)

    def test_restricted_items_from_confirmations(self):
        amounts = sorted(item.amount for item in self.ctx.lead_restricted)
        self.assertEqual(amounts, [200_000.0, 800_000.0])

    def test_recon_account_picks_primary_bank_account(self):
        info = self.ctx.recon_account
        self.assertIsNotNone(info)
        self.assertEqual(info.subject, "银行存款")
        self.assertEqual(info.bank_name, "招商银行")
        self.assertEqual(info.bank_account, "6225880200000001")
        self.assertEqual(self.ctx.recon_book_base, 5_432_109.0)
        self.assertEqual(self.ctx.recon_bank_base, 5_432_109.0)

    def test_recon_items_bucketed(self):
        buckets = self.ctx.recon_items_by_category
        self.assertEqual(len(buckets["book_plus"]), 1)
        self.assertEqual(len(buckets["bank_minus"]), 1)
        self.assertEqual(len(buckets["book_minus"]), 0)
        self.assertEqual(len(buckets["bank_plus"]), 0)
        self.assertAlmostEqual(buckets["book_plus"][0].amount, 500_000.0)
        self.assertAlmostEqual(buckets["bank_minus"][0].amount, 2_891.0)

    def test_cutoff_pre_and_post(self):
        self.assertEqual(len(self.ctx.cutoff_pre_period), 2)
        self.assertEqual(len(self.ctx.cutoff_post_period), 2)
        out_total = sum(sample.out_amount for sample in self.ctx.cutoff_pre_period)
        in_total = sum(sample.in_amount for sample in self.ctx.cutoff_pre_period)
        self.assertAlmostEqual(out_total + in_total, 0.0)

    def test_cutoff_window_default_used(self):
        self.assertEqual(self.ctx.cutoff_window, DEFAULT_CUTOFF_WINDOW)


class TestOptimizedTemplate(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.template = _find_template()
        cls.wb = load_workbook(cls.template, data_only=False)

    @classmethod
    def tearDownClass(cls):
        cls.wb.close()

    def test_five_visible_sheets_only(self):
        self.assertEqual(self.wb.sheetnames, list(clean.ALL_SHEETS))
        hidden = [ws.title for ws in self.wb.worksheets if ws.sheet_state != "visible"]
        self.assertEqual(hidden, [])

    def test_no_external_links(self):
        self.assertEqual(len(getattr(self.wb, "_external_links", [])), 0)

    def test_no_formula_error_literals(self):
        self.assertEqual(_formula_error_hits(self.wb), [])

    def test_threshold_formulas_preserved(self):
        lead = self.wb[clean.SHEET_LEAD]
        recon = self.wb[clean.SHEET_RECON]
        cutoff = self.wb[clean.SHEET_CUTOFF]
        self.assertTrue(str(lead["D15"].value).startswith("="))
        self.assertIn("MIN", str(recon["B3"].value))
        self.assertIn("MIN", str(cutoff["B3"].value))

    def test_package_has_no_forbidden_trace_tokens(self):
        forbidden = [
            bytes.fromhex(token)
            for token in (
                "4559",
                "e5ae89e6b0b8",
                "45726e7374",
                "596f756e67",
                "43616e766173",
                "47414d",
                "536b7977696e64",
                "4559496e7465727374617465",
                "563620535750",
            )
        ]
        hits = []
        with ZipFile(self.template) as package:
            for name in package.namelist():
                data = package.read(name)
                for token in forbidden:
                    if token in data:
                        hits.append((name, token))
        self.assertEqual(hits, [])


class TestCleanCellWriter(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.template = _find_template()
        cls.materials = load_case_materials(CASE_001_DIR)
        cls.ctx = build_fill_context(cls.materials)
        cls._tmp_ctx = writable_temp_dir()
        cls.tmp = cls._tmp_ctx.__enter__()
        cls.output = cls.tmp / "test_cash_filler_output.xlsx"
        write_cells_to_workbook(cls.template, cls.output, cls.ctx)
        cls.wb = load_workbook(cls.output, data_only=False)

    @classmethod
    def tearDownClass(cls):
        cls.wb.close()
        cls._tmp_ctx.__exit__(None, None, None)

    def test_lead_header_written(self):
        ws = self.wb[clean.SHEET_LEAD]
        self.assertEqual(ws[clean.LEAD_COMPANY_NAME].value, "星辰跨境科技(深圳)有限公司")
        period_end = ws[clean.LEAD_PERIOD_END].value
        if hasattr(period_end, "date"):
            period_end = period_end.date()
        self.assertEqual(period_end, date(2025, 12, 31))
        self.assertEqual(ws[clean.LEAD_TE].value, 5_000_000.0)
        self.assertEqual(ws[clean.LEAD_SAD].value, 250_000.0)
        self.assertEqual(ws[clean.LEAD_GAAP].value, "企业会计准则")
        self.assertEqual(ws[clean.LEAD_CURRENCY].value, "CNY")

    def test_clean_risk_levels_are_translated(self):
        ws = self.wb[clean.SHEET_LEAD]
        self.assertEqual(ws[clean.LEAD_RISK_COMPLETENESS].value, "中等")
        self.assertEqual(ws[clean.LEAD_RISK_EXISTENCE].value, "中等")
        self.assertEqual(ws[clean.LEAD_RISK_VALUATION].value, "低")

    def test_lead_balance_analysis_written(self):
        ws = self.wb[clean.SHEET_LEAD]
        self.assertEqual(ws["E29"].value, 32_500.0)
        self.assertEqual(ws["E30"].value, 7_545_565.0)
        self.assertEqual(ws["E31"].value, 800_000.0)
        self.assertTrue(str(ws["I33"].value).startswith("="))

    def test_bkd_rows_written(self):
        ws = self.wb[clean.SHEET_BKD]
        self.assertEqual(ws["C12"].value, "库存现金")
        self.assertEqual(ws["D13"].value, "招商银行")
        self.assertEqual(ws["E13"].value, "6225880200000001")
        self.assertEqual(ws["H13"].value, 5_432_109.0)
        self.assertEqual(ws["L15"].value, "银行承兑保证金")

    def test_bkd_tie_out_formulas_preserved(self):
        ws = self.wb[clean.SHEET_BKD]
        self.assertTrue(str(ws["H112"].value).startswith("="))
        self.assertTrue(str(ws["H114"].value).startswith("="))

    def test_recon_header_written(self):
        ws = self.wb[clean.SHEET_RECON]
        self.assertEqual(ws[clean.RECON_ACCOUNT_SUBJECT].value, "银行存款")
        self.assertEqual(ws[clean.RECON_BANK_NAME].value, "招商银行")
        self.assertEqual(ws[clean.RECON_BANK_ACCOUNT].value, "6225880200000001")
        self.assertEqual(ws[clean.RECON_CURRENCY].value, "CNY")
        self.assertEqual(ws[clean.RECON_BOOK_BASE].value, 5_432_109.0)
        self.assertEqual(ws[clean.RECON_BANK_BASE].value, 5_432_109.0)

    def test_recon_items_written(self):
        ws = self.wb[clean.SHEET_RECON]
        self.assertIn("已收未达", ws["A21"].value)
        self.assertEqual(ws["B21"].value, 500_000.0)
        self.assertEqual(ws["E27"].value, 2_891.0)
        self.assertTrue(str(ws["E33"].value).startswith("="))

    def test_cutoff_samples_written(self):
        ws = self.wb[clean.SHEET_CUTOFF]
        self.assertEqual(ws[clean.CUTOFF_WINDOW].value, DEFAULT_CUTOFF_WINDOW)
        self.assertEqual(ws["A13"].value, "S1")
        self.assertEqual(ws["A14"].value, "S2")
        self.assertEqual(ws["A38"].value, "S3")
        self.assertEqual(ws["A39"].value, "S4")

    def test_cutoff_check_formulas_preserved(self):
        ws = self.wb[clean.SHEET_CUTOFF]
        self.assertTrue(str(ws["P13"].value).startswith("="))
        self.assertTrue(str(ws["P38"].value).startswith("="))

    def test_filled_workbook_has_no_formula_error_literals(self):
        self.assertEqual(_formula_error_hits(self.wb), [])


class TestFillCashWorkpaperAPI(unittest.TestCase):
    def test_round_trip(self):
        template = _find_template()
        with writable_temp_dir() as tmp:
            output = tmp / "out.xlsx"
            result = fill_cash_workpaper(CASE_001_DIR, template, output)
            self.assertEqual(result, output)
            self.assertTrue(output.exists())
            self.assertGreater(output.stat().st_size, 20_000)


if __name__ == "__main__":
    unittest.main()
