# -*- coding: utf-8 -*-
"""Build the neutral optimized C cash workpaper template.

The output keeps the concise five-sheet structure of the simplified template
while adding enough professional structure for audit workpaper use:
risk-threshold formulas, tie-out checks, input/formula coloring, freeze panes,
and consistent section/table styling.
"""

from __future__ import annotations

import zipfile
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[2]
TEMPLATE_DIR = ROOT / "outputs" / "clean_templates"
BASE_TEMPLATE = TEMPLATE_DIR / "C_货币资金审计底稿_精简版.xlsx"
OUTPUT_TEMPLATE = TEMPLATE_DIR / "C_货币资金审计底稿_核心优化版_CN_CAS.xlsx"

BAD_TOKENS = [
    bytes.fromhex(token)
    for token in (
        "4559",
        "e5ae89e6b0b8",
        "45726e7374",
        "596f756e67",
        "43616e766173",
        "47414d",
        "536b7977696e64",
        "4559496e7465727374617465",
        "563620535750",
    )
]

FONT_NAME = "Microsoft YaHei"
COLORS = {
    "title": "263238",
    "section": "E7ECE7",
    "header": "4F6258",
    "subheader": "D9E5DD",
    "input": "FFF4CC",
    "formula": "EAF4F4",
    "check": "E9F5E8",
    "warning": "FCE7E7",
    "white": "FFFFFF",
    "grid": "D5DCE3",
    "text": "1F2933",
    "muted": "667085",
}

THIN = Side(style="thin", color=COLORS["grid"])
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def build_template() -> Path:
    TEMPLATE_DIR.mkdir(parents=True, exist_ok=True)
    if BASE_TEMPLATE.exists():
        wb = load_workbook(BASE_TEMPLATE)
        for ws in list(wb.worksheets):
            if ws.title not in {"汇总", "货币资金主表", "货币资金明细", "银行余额调节", "截止性测试"}:
                wb.remove(ws)
    else:
        wb = Workbook()
        wb.remove(wb.active)
        for name in ["汇总", "货币资金主表", "货币资金明细", "银行余额调节", "截止性测试"]:
            wb.create_sheet(name)

    _ensure_sheet_order(wb)
    for ws in wb.worksheets:
        _reset_sheet(ws)

    _build_summary(wb["汇总"])
    _build_lead(wb["货币资金主表"])
    _build_bkd(wb["货币资金明细"])
    _build_recon(wb["银行余额调节"])
    _build_cutoff(wb["截止性测试"])

    wb.active = 0
    wb.properties.creator = ""
    wb.properties.lastModifiedBy = ""
    wb.properties.title = "C Cash and Bank Audit Workpaper - Optimized Core"
    wb.properties.subject = "Brand-neutral clean cash and bank audit workpaper template"
    wb.properties.description = "Optimized five-sheet cash and bank audit workpaper template under CN CAS."
    wb.properties.keywords = "cash bank audit workpaper clean core optimized"
    wb.save(OUTPUT_TEMPLATE)
    _assert_clean_package(OUTPUT_TEMPLATE)
    return OUTPUT_TEMPLATE


def _ensure_sheet_order(wb) -> None:
    desired = ["汇总", "货币资金主表", "货币资金明细", "银行余额调节", "截止性测试"]
    for name in desired:
        if name not in wb.sheetnames:
            wb.create_sheet(name)
    wb._sheets = [wb[name] for name in desired]


def _reset_sheet(ws) -> None:
    for merged in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged))
    if ws.max_row:
        ws.delete_rows(1, ws.max_row)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = None
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.35
    ws.page_margins.right = 0.35
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5


def _build_summary(ws) -> None:
    ws.title = "汇总"
    _title(ws, "A1:J1", "货币资金审计底稿 - 核心优化版")
    _values(
        ws,
        "A3:B7",
        [
            ["客户名称", '=IF(货币资金主表!B3="","",货币资金主表!B3)'],
            ["期末日", '=IF(货币资金主表!B4="","",货币资金主表!B4)'],
            ["可容忍误差", '=IF(货币资金主表!B6="","",货币资金主表!B6)'],
            ["名义金额", '=IF(货币资金主表!B7="","",货币资金主表!B7)'],
            ["记账本位币", '=IF(货币资金主表!B10="","",货币资金主表!B10)'],
        ],
    )
    _section(ws, "A9:J9", "程序索引")
    _values(
        ws,
        "A10:J14",
        [
            ["编号", "名称", "主要认定", "工作目标", "执行状态", "编制人", "复核人", "链接/索引", "备注", "核对"],
            ["1", "货币资金主表", "E / V / C", "汇总风险评估、阈值、余额波动和披露核对。", "执行", "", "", "货币资金主表", "", '=IF(货币资金主表!B36="通过","通过","待核对")'],
            ["2", "货币资金明细", "E / V / C", "列示全量货币资金账户，并与主表期末余额核对。", "执行", "", "", "货币资金明细", "", '=IF(货币资金明细!B114="通过","通过","待核对")'],
            ["3", "银行余额调节", "E / C", "测试银行余额调节表及未达账项处理。", "执行", "", "", "银行余额调节", "", '=IF(银行余额调节!E33="通过","通过","待核对")'],
            ["4", "截止性测试", "E / C", "测试期末前后银行间转账是否记录于恰当期间。", "执行", "", "", "截止性测试", "", '=IF(COUNTIF(截止性测试!P13:P57,"<>0")=0,"通过","关注")'],
        ],
    )
    _section(ws, "A17:J17", "编制与复核")
    _values(ws, "A18:J20", [["编制人", "", "编制日期", "", "复核人", "", "复核日期", "", "总体备注", ""], ["", "", "", "", "", "", "", "", "", ""], ["说明", "黄色区域为输入区；浅蓝区域为公式区；绿色状态为核对通过。", "", "", "", "", "", "", "", ""]])
    _style_common(ws, max_row=22, max_col=10)
    _header(ws, "A10:J10")
    _input(ws, "B18:B19,D18:D19,F18:F19,H18:H19,J18:J19")
    _formula(ws, "B3:B7,J11:J14")
    _status_fill(ws, "E11:E14,J11:J14")
    ws.freeze_panes = "A10"
    _widths(ws, {"A": 12, "B": 18, "C": 12, "D": 46, "E": 12, "F": 12, "G": 12, "H": 18, "I": 18, "J": 12})
    ws.sheet_properties.tabColor = "4F6258"


def _build_lead(ws) -> None:
    _title(ws, "A1:N1", "货币资金主表")
    _values(
        ws,
        "A3:E10",
        [
            ["客户名称", "", "", "工作目标", "记录货币资金余额、账户完整性、重大波动和后续程序结论"],
            ["期末日", "", "", "填写说明", "黄色区域为输入区；浅蓝区域为公式或自动检查区"],
            ["分析日期", "", "", "", ""],
            ["可容忍误差", "", "", "", ""],
            ["名义金额", "", "", "", ""],
            ["适用会计准则", "中国企业会计准则", "", "", ""],
            ["审计准则", "中国注册会计师审计准则", "", "", ""],
            ["记账本位币", "CNY", "", "", ""],
        ],
    )
    _section(ws, "A13:N13", "认定风险与测试阈值")
    _values(
        ws,
        "A14:E19",
        [
            ["认定", "CRA 风险等级", "阈值比例", "测试阈值", "说明"],
            ["完整性", "", '=IF(B15="不适用","",IF(B15="极低",1,IF(B15="低",0.75,IF(B15="中等",0.5,IF(B15="高",0.25,"")))))', '=IF(C15="","",$B$6*C15)', "C.02 / 截止测试取完整性与存在性阈值较小值"],
            ["存在性", "", '=IF(B16="不适用","",IF(B16="极低",1,IF(B16="低",0.75,IF(B16="中等",0.5,IF(B16="高",0.25,"")))))', '=IF(C16="","",$B$6*C16)', ""],
            ["准确性/计价", "", '=IF(B17="不适用","",IF(B17="极低",1,IF(B17="低",0.75,IF(B17="中等",0.5,IF(B17="高",0.25,"")))))', '=IF(C17="","",$B$6*C17)', ""],
            ["权利和义务", "", '=IF(B18="不适用","",IF(B18="极低",1,IF(B18="低",0.75,IF(B18="中等",0.5,IF(B18="高",0.25,"")))))', '=IF(C18="","",$B$6*C18)', ""],
            ["列报", "", '=IF(B19="不适用","",IF(B19="极低",1,IF(B19="低",0.75,IF(B19="中等",0.5,IF(B19="高",0.25,"")))))', '=IF(C19="","",$B$6*C19)', ""],
        ],
    )
    _values(ws, "A23:D23", [["波动比例阈值", "", "余额波动同时参考名义金额与比例阈值", ""]])
    _section(ws, "A27:N27", "余额波动分析")
    _values(
        ws,
        "A28:N33",
        [
            ["账套", "科目编码", "科目", "索引", "期末账面数", "账表调整", "期末未审数", "审计调整", "期末审定数", "上年审定数", "变动金额", "变动%", "超阈值", "说明"],
            ["", "", "库存现金", "货币资金明细", "", "", '=IF(OR(E29<>"",F29<>""),SUM(E29:F29),"")', "", '=IF(OR(G29<>"",H29<>""),SUM(G29:H29),"")', "", '=IF(OR(I29<>"",J29<>""),I29-J29,"")', '=IF(J29<>0,K29/J29,IF(I29<>0,1,0))', '=IF(OR(K29="",$B$7="",$B$23=""),"",IF(OR(ABS(K29)>=$B$7,ABS(L29)>=$B$23),"是","否"))', ""],
            ["", "", "银行存款", "货币资金明细", "", "", '=IF(OR(E30<>"",F30<>""),SUM(E30:F30),"")', "", '=IF(OR(G30<>"",H30<>""),SUM(G30:H30),"")', "", '=IF(OR(I30<>"",J30<>""),I30-J30,"")', '=IF(J30<>0,K30/J30,IF(I30<>0,1,0))', '=IF(OR(K30="",$B$7="",$B$23=""),"",IF(OR(ABS(K30)>=$B$7,ABS(L30)>=$B$23),"是","否"))', ""],
            ["", "", "其他货币资金", "货币资金明细", "", "", '=IF(OR(E31<>"",F31<>""),SUM(E31:F31),"")', "", '=IF(OR(G31<>"",H31<>""),SUM(G31:H31),"")', "", '=IF(OR(I31<>"",J31<>""),I31-J31,"")', '=IF(J31<>0,K31/J31,IF(I31<>0,1,0))', '=IF(OR(K31="",$B$7="",$B$23=""),"",IF(OR(ABS(K31)>=$B$7,ABS(L31)>=$B$23),"是","否"))', ""],
            ["", "", "", "", "", "", "", "", "", "", "", "", "", ""],
            ["合计", "", "", "", '=SUM(E29:E31)', '=SUM(F29:F31)', '=SUM(G29:G31)', '=SUM(H29:H31)', '=SUM(I29:I31)', '=SUM(J29:J31)', '=I33-J33', '=IF(J33<>0,K33/J33,IF(I33<>0,1,0))', '=IF(OR(K33="",$B$7="",$B$23=""),"",IF(OR(ABS(K33)>=$B$7,ABS(L33)>=$B$23),"是","否"))', ""],
        ],
    )
    _values(ws, "A35:B36", [["与货币资金明细期末本币金额差异", '=货币资金明细!H112-I33'], ["核对状态", '=IF(ROUND(B35,2)=0,"通过","待核对")']])
    _section(ws, "A42:N42", "受限货币资金披露")
    _values(ws, "A44:F45", [["项目", "期末审定数", "上年审定数", "差异", "核对状态", "说明"], ["所有权或使用权受限货币资金", "", "", '=IF(OR(B45<>"",C45<>""),B45-C45,"")', '=IF(ROUND(B45,2)=0,"无受限资金","关注披露")', ""]])
    _style_common(ws, max_row=50, max_col=14)
    _header(ws, "A14:E14,A28:N28,A44:F44")
    _input(ws, "B3:B10,B15:B19,B23,E29:F31,H29:H31,J29:J31,N29:N31,B45:C45,F45")
    _formula(ws, "C15:D19,G29:G31,I29:I31,K29:M33,E33:M33,B35:B36,D45:E45")
    _status_fill(ws, "B36,E45")
    _add_list_validation(ws, "B15:B19", ["极低", "低", "中等", "高", "不适用"])
    ws.freeze_panes = "A28"
    _widths(ws, {"A": 16, "B": 14, "C": 16, "D": 18, "E": 15, "F": 13, "G": 15, "H": 13, "I": 15, "J": 15, "K": 15, "L": 12, "M": 12, "N": 28})
    ws.sheet_properties.tabColor = "6B7D6D"


def _build_bkd(ws) -> None:
    _title(ws, "A1:L1", "货币资金明细")
    _values(ws, "A3:L4", [["列示所有货币资金账户，核对至科目余额表", "", "", "", "", "", "", "", "", "", "", ""], ["资料来源：科目余额表 / 函证或银行回函 / 银行流水", "", "", "", "", "", "", "", "", "", "", ""]])
    _values(ws, "A10:L10", [["序号", "公司", "科目", "银行", "账号", "币种", "期末金额(原币)", "期末金额(本币)", "用途", "是否函证", "替代程序", "备注"]])
    for row in range(12, 112):
        ws[f"A{row}"] = f'=IF(COUNTA(B{row}:L{row})=0,"",ROW()-11)'
    _values(ws, "A112:L114", [["合计", "", "", "", "", "", '=SUM(G12:G111)', '=SUM(H12:H111)', "", "", "", ""], ["与主表期末审定数差异", "", "", "", "", "", "", '=H112-货币资金主表!I33', "", "", "", ""], ["核对状态", "", "", "", "", "", "", '=IF(ROUND(H113,2)=0,"通过","待核对")', "", "", "", ""]])
    _style_common(ws, max_row=114, max_col=12)
    _header(ws, "A10:L10")
    _input(ws, "B12:L111")
    _formula(ws, "A12:A111,G112:H114")
    _status_fill(ws, "H114")
    _add_list_validation(ws, "F12:F111", ["CNY", "USD", "EUR", "HKD", "JPY", "GBP"])
    _add_list_validation(ws, "J12:J111", ["是", "否"])
    ws.freeze_panes = "A12"
    _widths(ws, {"A": 8, "B": 22, "C": 16, "D": 20, "E": 22, "F": 10, "G": 16, "H": 16, "I": 22, "J": 12, "K": 18, "L": 24})
    ws.sheet_properties.tabColor = "7A8F7B"


def _build_recon(ws) -> None:
    _title(ws, "A1:G1", "银行余额调节")
    _values(
        ws,
        "A3:G6",
        [
            ["测试阈值", '=IF(COUNT(货币资金主表!D15:D16)=0,"",MIN(货币资金主表!D15:D16))', "用于识别和测试调节项", "", "", "", ""],
            ["截止期间", "", "由项目组根据风险评估、账户性质和交易处理周期确定", "", "", "", ""],
            ["选取说明", "", "", "", "", "", ""],
            ["结论", "", "", "", "", "", ""],
        ],
    )
    _section(ws, "A9:G9", "账户信息")
    _values(ws, "A10:G12", [["科目", "", "索引", "", "编制日期", "", ""], ["银行", "", "账号", "", "币种", "", ""], ["账户性质", "", "有无调节项", "", "资料", "", ""]])
    _section(ws, "A17:G17", "余额调节表")
    _values(
        ws,
        "A18:G33",
        [
            ["账面侧", "账面金额(原币)", "索引", "银行侧", "对账单金额", "索引", "备注"],
            ["期末账面余额", "", "", "银行对账单余额", "", "", ""],
            ["加：银收企未收", "=SUM(B21:B25)", "", "加：企收银未收", "=SUM(E21:E25)", "", ""],
            ["明细1", "", "", "明细1", "", "", ""],
            ["明细2", "", "", "明细2", "", "", ""],
            ["明细3", "", "", "明细3", "", "", ""],
            ["明细4", "", "", "明细4", "", "", ""],
            ["明细5", "", "", "明细5", "", "", ""],
            ["减：银付企未付", "=SUM(B27:B31)", "", "减：企付银未付", "=SUM(E27:E31)", "", ""],
            ["明细1", "", "", "明细1", "", "", ""],
            ["明细2", "", "", "明细2", "", "", ""],
            ["明细3", "", "", "明细3", "", "", ""],
            ["明细4", "", "", "明细4", "", "", ""],
            ["明细5", "", "", "明细5", "", "", ""],
            ["调节后金额", "=B19+B20-B26", "", "调节后金额", "=E19+E20-E26", "", ""],
            ["核对", "=E32-B32", "", "核对状态", '=IF(ROUND(B33,2)=0,"通过","待核对")', "", ""],
        ],
    )
    _style_common(ws, max_row=36, max_col=7)
    _header(ws, "A18:G18")
    _input(ws, "B4:B6,B10:B12,D10:D12,F10:F12,B19:B31,E19:E31,C21:C31,F21:F31")
    _formula(ws, "B3,B20:B33,E20:E33")
    _status_fill(ws, "E33")
    _add_list_validation(ws, "D12:D12", ["是", "否"])
    ws.freeze_panes = "A18"
    _widths(ws, {"A": 20, "B": 16, "C": 16, "D": 20, "E": 16, "F": 16, "G": 22})
    ws.sheet_properties.tabColor = "8C9A70"


def _build_cutoff(ws) -> None:
    _title(ws, "A1:P1", "截止性测试")
    _values(
        ws,
        "A3:P7",
        [
            ["测试阈值", '=IF(COUNT(货币资金主表!D15:D16)=0,"",MIN(货币资金主表!D15:D16))', "用于识别和测试银行间转账", "", "", "", "", "", "", "", "", "", "", "", "", ""],
            ["截止期间", "", "例如期末日前后 1 个工作日；可按风险评估延长", "", "", "", "", "", "", "", "", "", "", "", "", ""],
            ["使用的截止期间（工作日）", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
            ["选取理由", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
            ["结论", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
        ],
    )
    _section(ws, "A11:P11", "期末前银行间转账")
    cutoff_headers = ["样本编号", "付款/收款主体", "转出银行", "转出账号", "付款/收款时间", "交易编号", "币种", "转出金额", "转入银行", "转入账号", "入账日期", "转入金额", "跨期是否体现在余额调节表", "会计处理是否恰当", "备注", "核对"]
    _values(ws, "A12:P12", [cutoff_headers])
    for row in range(13, 33):
        ws[f"P{row}"] = f'=IF(OR(H{row}<>"",L{row}<>""),H{row}+L{row},"")'
    _section(ws, "A36:P36", "期末后银行间转账")
    _values(ws, "A37:P37", [cutoff_headers])
    for row in range(38, 58):
        ws[f"P{row}"] = f'=IF(OR(H{row}<>"",L{row}<>""),H{row}+L{row},"")'
    _values(ws, "A60:P61", [["核对汇总", "", "", "", "", "", "", "", "", "", "", "", "", "", "异常样本数", '=COUNTIF(P13:P57,"<>0")'], ["核对状态", "", "", "", "", "", "", "", "", "", "", "", "", "", "", '=IF(P60=0,"通过","关注")']])
    _style_common(ws, max_row=62, max_col=16)
    _header(ws, "A12:P12,A37:P37")
    _input(ws, "B4:B7,A13:O32,A38:O57")
    _formula(ws, "B3,P13:P32,P38:P61")
    _status_fill(ws, "P61")
    _add_list_validation(ws, "G13:G57", ["CNY", "USD", "EUR", "HKD", "JPY", "GBP"])
    _add_list_validation(ws, "M13:N57", ["是", "否", "不适用"])
    ws.freeze_panes = "A12"
    _widths(ws, {"A": 12, "B": 22, "C": 18, "D": 20, "E": 16, "F": 18, "G": 10, "H": 15, "I": 18, "J": 20, "K": 14, "L": 15, "M": 22, "N": 18, "O": 22, "P": 12})
    ws.sheet_properties.tabColor = "A09363"


def _title(ws, range_addr: str, value: str) -> None:
    ws.merge_cells(range_addr)
    cell = ws[range_addr.split(":")[0]]
    cell.value = value
    cell.fill = PatternFill("solid", fgColor=COLORS["title"])
    cell.font = Font(name=FONT_NAME, bold=True, color=COLORS["white"], size=16)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[cell.row].height = 28


def _section(ws, range_addr: str, value: str) -> None:
    ws.merge_cells(range_addr)
    cell = ws[range_addr.split(":")[0]]
    cell.value = value
    cell.fill = PatternFill("solid", fgColor=COLORS["section"])
    cell.font = Font(name=FONT_NAME, bold=True, color=COLORS["text"], size=11)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def _values(ws, range_addr: str, matrix: list[list[object]]) -> None:
    start, end = range_addr.split(":")
    start_col, start_row = _split_cell(start)
    end_col, end_row = _split_cell(end)
    for r_idx, row in enumerate(range(start_row, end_row + 1)):
        for c_idx, col in enumerate(range(start_col, end_col + 1)):
            value = matrix[r_idx][c_idx] if r_idx < len(matrix) and c_idx < len(matrix[r_idx]) else None
            ws.cell(row, col).value = value


def _style_common(ws, max_row: int, max_col: int) -> None:
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.font = Font(name=FONT_NAME, size=10, color=COLORS["text"])
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = BORDER
    for row_idx in range(1, max_row + 1):
        ws.row_dimensions[row_idx].height = 22
    for cell in ws[1]:
        cell.font = Font(name=FONT_NAME, bold=True, color=COLORS["white"], size=16)


def _header(ws, ranges: str) -> None:
    for addr in ranges.split(","):
        for row in _range_rows(ws, addr):
            for cell in row:
                cell.fill = PatternFill("solid", fgColor=COLORS["header"])
                cell.font = Font(name=FONT_NAME, bold=True, color=COLORS["white"], size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _input(ws, ranges: str) -> None:
    for addr in ranges.split(","):
        for row in _range_rows(ws, addr):
            for cell in row:
                cell.fill = PatternFill("solid", fgColor=COLORS["input"])


def _formula(ws, ranges: str) -> None:
    for addr in ranges.split(","):
        for row in _range_rows(ws, addr):
            for cell in row:
                cell.fill = PatternFill("solid", fgColor=COLORS["formula"])


def _status_fill(ws, ranges: str) -> None:
    for addr in ranges.split(","):
        for row in _range_rows(ws, addr):
            for cell in row:
                cell.fill = PatternFill("solid", fgColor=COLORS["check"])
                ws.conditional_formatting.add(
                    cell.coordinate,
                    CellIsRule(operator="equal", formula=['"待核对"'], fill=PatternFill("solid", fgColor=COLORS["warning"])),
                )


def _add_list_validation(ws, sqref: str, values: list[str]) -> None:
    dv = DataValidation(type="list", formula1=f'"{",".join(values)}"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(sqref)


def _widths(ws, widths: dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def _split_cell(addr: str) -> tuple[int, int]:
    from openpyxl.utils.cell import column_index_from_string, coordinate_from_string

    col, row = coordinate_from_string(addr)
    return column_index_from_string(col), row


def _range_rows(ws, addr: str):
    obj = ws[addr]
    if isinstance(obj, tuple):
        if obj and isinstance(obj[0], tuple):
            return obj
        return (obj,)
    return ((obj,),)


def _assert_clean_package(path: Path) -> None:
    with zipfile.ZipFile(path) as archive:
        hits = []
        for name in archive.namelist():
            data = archive.read(name)
            for token in BAD_TOKENS:
                if token in data:
                    hits.append((name, token.decode("utf-8", "ignore")))
        if hits:
            raise RuntimeError(f"Forbidden workbook traces found: {hits[:10]}")


if __name__ == "__main__":
    print(build_template())
