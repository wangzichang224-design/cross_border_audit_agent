# -*- coding: utf-8 -*-
"""
Agent core: fill the C 货币资金 audit workpaper from a materials directory.

End-to-end flow
---------------
::

    materials_dir/  ──[materials_loader]──>  CaseMaterials
                                                  │
                                          [build_fill_context]
                                                  │
                                                  v
                                            FillContext
                                                  │
                                       [write_cells_to_workbook]
                                                  │
                                                  v
                                       output_path/*.xlsx (filled)

In Phase 1 the "audit reasoning" in `build_fill_context` is **rule-based
and minimal**:
- Aggregate the period-end balance for each fixed account
  (库存现金 / 银行存款 / 其他货币资金) directly from period_summary.csv
- Use the first confirmation as the C.02 reconciliation account header
- Carry reconciliation_items.csv straight into the 4 category buckets
- Pick inter-bank transfers from bank_statement.csv near period_end as
  cutoff samples

Phase 3 swaps `materials_loader` for PDF parsers AND replaces
`build_fill_context` with an LLM-guided agent. The cell writer below
(`write_cells_to_workbook`) stays unchanged across phases — it's a
pure mechanical translator from FillContext to xlsx cells.
"""

from __future__ import annotations

import logging
import shutil
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

from . import cell_map as cm
from . import cell_map_clean as cm_clean
from .materials_loader import (
    CaseMaterials,
    PeriodSummaryRow,
    ReconciliationItem,
    load_case_materials,
)


logger = logging.getLogger(__name__)

CLEAN_SHEET_ALIASES = {
    "summary": (cm_clean.SHEET_SUMMARY,),
    "lead": (cm_clean.SHEET_LEAD, "C.00 Lead"),
    "bkd": (cm_clean.SHEET_BKD, "C.00 BKD"),
    "recon": (cm_clean.SHEET_RECON, "C.02 银行余额调节"),
    "cutoff": (cm_clean.SHEET_CUTOFF, "C.03 截止性测试"),
}


# The three fixed GL-account names that appear in C.00 Lead rows 38-40.
# Tied to cell_map.LEAD_GL_ROWS — order matters for downstream tests.
FIXED_GL_ACCOUNTS = ("库存现金", "银行存款", "其他货币资金")

# Phase 1 default risk levels — conservative defaults until an audit plan
# is actually attached to the case. Override by setting these in
# CaseMetadata in future phases.
DEFAULT_RISK_LEVELS = {
    cm.LEAD_RISK_COMPLETENESS: "Moderate",
    cm.LEAD_RISK_EXISTENCE: "Moderate",
    cm.LEAD_RISK_VALUATION: "Low",
    cm.LEAD_RISK_RIGHTS: "Low",
    cm.LEAD_RISK_PRESENTATION: "Low",
}

# Default cutoff testing window text (matches what the EY template expects)
DEFAULT_CUTOFF_WINDOW = "期末前后 5 个工作日"


# ──────────────────────────────────────────────────────────────────────────────
# FillContext — what gets passed to the cell writer
# ──────────────────────────────────────────────────────────────────────────────


@dataclass(frozen=True)
class ReconAccountInfo:
    subject: str            # 科目 (e.g. "银行存款")
    bank_name: str
    bank_account: str
    currency: str
    statement_date: date


@dataclass(frozen=True)
class FillContext:
    """Pre-computed values, ready for cell writing.

    Each field maps to a section of one of the three sheets.  The cell
    writer treats this as an immutable bundle of facts; all reasoning
    has already happened upstream.
    """

    # ── C.00 Lead ────────────────────────────────────────────────
    company_name: str
    period_end: date
    analysis_date: date
    te: float
    sad: float
    gaap: str
    currency: str
    variation_pct: float
    risk_levels: dict[str, str]                       # cell_addr -> level
    lead_gl_rows: dict[str, cm.LeadGLRow]             # account_name -> row
    lead_disclosure: dict[str, tuple[float, float]]   # account_name -> (current, prior)
    lead_restricted: list[cm.LeadRestrictedItem]

    # ── C.02 Bank reconciliations ────────────────────────────────
    recon_account: ReconAccountInfo | None
    recon_book_base: float
    recon_bank_base: float
    recon_items_by_category: dict[str, list[cm.ReconItem]]
    # keys: 'book_plus' / 'book_minus' / 'bank_plus' / 'bank_minus'

    # ── C.03 Cutoff ──────────────────────────────────────────────
    cutoff_window: str
    cutoff_pre_period: list[cm.CutoffSample]
    cutoff_post_period: list[cm.CutoffSample]

    # Clean-template support: full account-level rows for C.00 BKD.
    period_summary_rows: tuple[PeriodSummaryRow, ...] = field(default_factory=tuple)
    lead_gl_notes: dict[str, str] = field(default_factory=dict)
    recon_rationale: str = ""
    recon_conclusion: str = ""
    cutoff_reason: str = ""
    cutoff_conclusion: str = ""


# ──────────────────────────────────────────────────────────────────────────────
# Context builder — "minimal audit reasoning" for Phase 1
# ──────────────────────────────────────────────────────────────────────────────


def build_fill_context(materials: CaseMaterials) -> FillContext:
    """Translate parsed CSV materials into a ready-to-write FillContext.

    Phase 1 logic only — does NOT detect anomalies. Subsequent phases will
    layer rule-based and LLM-driven audit reasoning on top of this baseline.
    """
    meta = materials.meta

    lead_gl_rows = _build_lead_gl_rows(materials)
    lead_disclosure = _build_lead_disclosure(materials, lead_gl_rows)
    lead_restricted = _build_lead_restricted(materials)
    recon_account, recon_book_base, recon_bank_base = _build_recon_header(materials)
    recon_items = _bucket_recon_items(materials.reconciliation_items)
    cutoff_pre, cutoff_post = _build_cutoff_samples(materials)

    return FillContext(
        company_name=meta.client_name,
        period_end=meta.period_end,
        analysis_date=meta.analysis_date,
        te=meta.te,
        sad=meta.sad,
        gaap=meta.gaap,
        currency=meta.currency,
        variation_pct=meta.variation_pct,
        risk_levels=DEFAULT_RISK_LEVELS,
        lead_gl_rows=lead_gl_rows,
        lead_disclosure=lead_disclosure,
        lead_restricted=lead_restricted,
        recon_account=recon_account,
        recon_book_base=recon_book_base,
        recon_bank_base=recon_bank_base,
        recon_items_by_category=recon_items,
        cutoff_window=DEFAULT_CUTOFF_WINDOW,
        cutoff_pre_period=cutoff_pre,
        cutoff_post_period=cutoff_post,
        period_summary_rows=materials.period_summary,
    )


def _build_lead_gl_rows(materials: CaseMaterials) -> dict[str, cm.LeadGLRow]:
    """Aggregate period_summary by account_name into LeadGLRow values.

    Returns one row per fixed account, even if no period_summary rows
    exist for that account (in which case all amounts are zero).
    """
    by_account: dict[str, list[PeriodSummaryRow]] = {a: [] for a in FIXED_GL_ACCOUNTS}
    for row in materials.period_summary:
        if row.account_name in by_account:
            by_account[row.account_name].append(row)
        else:
            logger.warning(
                "period_summary row with unmapped account_name %r — ignored",
                row.account_name,
            )

    out: dict[str, cm.LeadGLRow] = {}
    for account_name, rows in by_account.items():
        current = sum(r.period_end_balance_local for r in rows)
        prior = sum(r.prior_year_balance for r in rows)
        out[account_name] = cm.LeadGLRow(
            gl_account_code=_default_gl_code(account_name),
            index_ref="C.00 BKD/",
            book_value_unaudited=current,
            prior_year_audited=prior,
        )
    return out


def _default_gl_code(account_name: str) -> str:
    """Standard CAS GL codes for the three cash accounts."""
    return {
        "库存现金": "1001",
        "银行存款": "1002",
        "其他货币资金": "1009",
    }.get(account_name, "")


def _build_lead_disclosure(
    materials: CaseMaterials,
    gl_rows: dict[str, cm.LeadGLRow],
) -> dict[str, tuple[float, float]]:
    """Build the C.00 Lead 表2 disclosure values from GL aggregation.

    For now: 表2 mirrors 表1 (库存现金/银行存款/其他货币资金) and adds
    a "存放财务公司款项" row that defaults to (0, 0) — finance-company
    deposits are uncommon in cross-border e-commerce.
    """
    out: dict[str, tuple[float, float]] = {}
    for account_name in FIXED_GL_ACCOUNTS:
        row = gl_rows.get(account_name)
        if row:
            out[account_name] = (row.book_value_unaudited, row.prior_year_audited)
        else:
            out[account_name] = (0.0, 0.0)
    out["存放财务公司款项"] = (0.0, 0.0)
    return out


def _build_lead_restricted(materials: CaseMaterials) -> list[cm.LeadRestrictedItem]:
    """Pull restricted-cash items.

    Source priority:
      1. Confirmation footnotes (restricted_amount > 0)
      2. period_summary rows flagged is_restricted=True (fallback;
         confirmations may not always be available)

    This is also the FIRST place the Agent could spot a classification
    error: if a confirmation says "质押冻结 200,000" but period_summary
    says is_restricted=False, the Agent's view diverges from the client's.
    For Phase 1 we just emit whatever the confirmations say.
    """
    items: list[cm.LeadRestrictedItem] = []
    seen_accounts: set[str] = set()
    for c in materials.confirmations:
        if c.restricted_amount > 0:
            items.append(
                cm.LeadRestrictedItem(
                    description=f"{c.bank_name} {c.bank_account[-4:]} 受限",
                    amount=c.restricted_amount,
                    nature=c.restriction_nature or "受限",
                    index=c.confirmation_index or "C.01",
                )
            )
            seen_accounts.add(c.bank_account)
    # Fallback to period_summary for accounts without a matching confirmation
    for p in materials.period_summary:
        if p.is_restricted and p.bank_account not in seen_accounts:
            items.append(
                cm.LeadRestrictedItem(
                    description=f"{p.bank_name} {p.bank_account[-4:]} 受限",
                    amount=p.period_end_balance_local,
                    nature=p.restriction_note or "受限",
                    index="C.00b",
                )
            )
    # Respect the cell-map capacity
    return items[: cm.MAX_RESTRICTED_ITEMS]


def _build_recon_header(
    materials: CaseMaterials,
) -> tuple[ReconAccountInfo | None, float, float]:
    """Pick the first 银行存款 account that has BOTH a confirmation and a
    period_summary row as the C.02 reconciliation target.

    Returns (account_info, book_base, bank_base). Returns (None, 0, 0)
    if no eligible account exists, which is OK — the writer will just
    leave the reconciliation section blank.
    """
    if not materials.confirmations or not materials.period_summary:
        return None, 0.0, 0.0

    # Prefer 银行存款; fall back to first confirmation.
    bank_accounts = {c.bank_account: c for c in materials.confirmations}
    candidate = None
    for ps in materials.period_summary:
        if ps.account_name == "银行存款" and ps.bank_account in bank_accounts:
            candidate = (ps, bank_accounts[ps.bank_account])
            break
    if candidate is None:
        # Fallback: any matching account
        for ps in materials.period_summary:
            if ps.bank_account in bank_accounts:
                candidate = (ps, bank_accounts[ps.bank_account])
                break
    if candidate is None:
        return None, 0.0, 0.0

    ps, conf = candidate
    info = ReconAccountInfo(
        subject=ps.account_name,
        bank_name=ps.bank_name or conf.bank_name,
        bank_account=ps.bank_account,
        currency=ps.currency or conf.currency,
        statement_date=conf.confirmation_date or materials.meta.period_end,
    )
    # Book base = period_summary in original currency
    book_base = ps.period_end_balance_fx if ps.period_end_balance_fx else ps.period_end_balance_local
    # Bank base = confirmation amount
    bank_base = conf.confirmed_balance
    return info, book_base, bank_base


def _bucket_recon_items(
    recon_items: tuple[ReconciliationItem, ...],
) -> dict[str, list[cm.ReconItem]]:
    """Split reconciliation_items.csv into 4 buckets by category, respecting
    the per-category max of MAX_RECON_ITEMS_PER_CATEGORY.
    """
    buckets: dict[str, list[cm.ReconItem]] = {c: [] for c in (
        "book_plus", "book_minus", "bank_plus", "bank_minus",
    )}
    for item in recon_items:
        if len(buckets[item.category]) >= cm.MAX_RECON_ITEMS_PER_CATEGORY:
            logger.warning(
                "reconciliation category %s exceeded max %d; dropping %r",
                item.category, cm.MAX_RECON_ITEMS_PER_CATEGORY, item.description,
            )
            continue
        buckets[item.category].append(
            cm.ReconItem(
                description=item.description,
                amount=item.amount,
                index=item.index,
            )
        )
    return buckets


def _build_cutoff_samples(
    materials: CaseMaterials,
) -> tuple[list[cm.CutoffSample], list[cm.CutoffSample]]:
    """Identify inter-bank transfer candidates around period_end.

    Heuristic for "inter-bank transfer":
      * Has a non-empty counterparty AND counterparty is itself a known
        bank account (i.e. the company's own other account)
      OR
      * Description contains an inter-account keyword like "转账"/"调拨"

    For Phase 1 the bank_statement.csv fixtures will pre-label these via
    the description field. Phase 3's smarter agent will infer membership
    based on the audit-entity's bank-account ledger.
    """
    period_end = materials.meta.period_end
    window = timedelta(days=7)  # ~5 business days
    pre_lo, pre_hi = period_end - window, period_end
    post_lo, post_hi = period_end + timedelta(days=1), period_end + window

    pre_samples: list[cm.CutoffSample] = []
    post_samples: list[cm.CutoffSample] = []

    own_accounts = {ps.bank_account for ps in materials.period_summary}

    for txn in materials.bank_statements:
        if not _is_inter_bank_transfer(txn, own_accounts):
            continue
        sample = _txn_to_cutoff_sample(txn, materials.meta.client_name, len(pre_samples) + len(post_samples) + 1)
        if pre_lo <= txn.date <= pre_hi and len(pre_samples) < cm.MAX_CUTOFF_SAMPLES_PER_TABLE:
            pre_samples.append(sample)
        elif post_lo <= txn.date <= post_hi and len(post_samples) < cm.MAX_CUTOFF_SAMPLES_PER_TABLE:
            post_samples.append(sample)

    return pre_samples, post_samples


def _is_inter_bank_transfer(txn, own_accounts: set[str]) -> bool:
    """Heuristic for whether a bank statement txn is an inter-bank transfer."""
    if txn.counterparty and txn.counterparty in own_accounts:
        return True
    if txn.description and any(kw in txn.description for kw in ("银行间转账", "调拨", "内部转账")):
        return True
    return False


def _txn_to_cutoff_sample(txn, company_name: str, sample_idx: int) -> cm.CutoffSample:
    out_amount = -abs(txn.debit) if txn.debit else 0.0  # outflow → negative
    in_amount = abs(txn.credit) if txn.credit else 0.0
    return cm.CutoffSample(
        sample_id=f"S{sample_idx}",
        company_name=company_name,
        out_bank_name=txn.bank_name if txn.debit else "",
        out_bank_account=txn.bank_account if txn.debit else "",
        out_time=txn.date.isoformat(),
        txn_id=txn.txn_id or f"T{sample_idx:04d}",
        currency=txn.currency,
        out_amount=out_amount,
        in_bank_name=txn.bank_name if txn.credit else (txn.counterparty or ""),
        in_bank_account=txn.counterparty if txn.credit else "",
        in_date=txn.date.isoformat(),
        in_amount=in_amount,
        is_in_reconciliation="是",
        notes="",
    )


# ──────────────────────────────────────────────────────────────────────────────
# Openpyxl cell writer
# ──────────────────────────────────────────────────────────────────────────────


def _detect_template_layout(wb) -> str:
    """Return ``clean`` or ``legacy`` based on workbook sheet names."""
    if all(_first_existing_sheet_name(wb, aliases) for aliases in CLEAN_SHEET_ALIASES.values()):
        return "clean"
    return "legacy"


def _first_existing_sheet_name(wb, aliases: tuple[str, ...]) -> str:
    for name in aliases:
        if name in wb.sheetnames:
            return name
    return ""


def _clean_sheet(wb, key: str):
    name = _first_existing_sheet_name(wb, CLEAN_SHEET_ALIASES[key])
    if not name:
        expected = " or ".join(CLEAN_SHEET_ALIASES[key])
        raise KeyError(f"Clean template sheet not found: {expected}")
    return wb[name]


def _set_cell(ws, addr: str, value) -> None:
    """Set a cell, redirecting merged-cell writes to the merged range anchor."""
    cell = ws[addr]
    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if addr in merged_range:
                ws.cell(merged_range.min_row, merged_range.min_col).value = value
                return
    cell.value = value


def write_cells_to_workbook(
    template_path: Path,
    output_path: Path,
    context: FillContext,
) -> Path:
    """Copy template → write all FillContext data → save to output_path.

    The original template is never modified. ``keep_vba`` is auto-set
    when the suffix is ``.xlsm`` so macro templates round-trip cleanly.
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template_path, output_path)
    keep_vba = output_path.suffix.lower() == ".xlsm"
    wb = load_workbook(output_path, keep_vba=keep_vba)

    layout = _detect_template_layout(wb)
    if layout == "clean":
        _write_clean_workbook(wb, context)
        wb.save(output_path)
        return output_path

    if cm.SHEET_LEAD in wb.sheetnames:
        _write_lead_sheet(wb[cm.SHEET_LEAD], context)
    else:
        logger.warning("Sheet %r not found in template; skipping", cm.SHEET_LEAD)

    if cm.SHEET_RECON in wb.sheetnames:
        _write_recon_sheet(wb[cm.SHEET_RECON], context)
    else:
        logger.warning("Sheet %r not found in template; skipping", cm.SHEET_RECON)

    if cm.SHEET_CUTOFF in wb.sheetnames:
        _write_cutoff_sheet(wb[cm.SHEET_CUTOFF], context)
    else:
        logger.warning("Sheet %r not found in template; skipping", cm.SHEET_CUTOFF)

    wb.save(output_path)
    return output_path


def _write_lead_sheet(ws, ctx: FillContext) -> None:
    """Write C.00 Lead — header, risk levels, GL table, disclosure, restricted."""
    # Header block
    ws[cm.LEAD_COMPANY_NAME] = ctx.company_name
    ws[cm.LEAD_PERIOD_END] = ctx.period_end
    ws[cm.LEAD_ANALYSIS_DATE] = ctx.analysis_date
    ws[cm.LEAD_TE] = ctx.te
    ws[cm.LEAD_SAD] = ctx.sad
    ws[cm.LEAD_GAAP] = ctx.gaap
    ws[cm.LEAD_CURRENCY] = ctx.currency
    ws[cm.LEAD_VARIATION_PCT] = ctx.variation_pct

    # Risk levels (drives D-column threshold formulas)
    for addr, level in ctx.risk_levels.items():
        if level not in cm.ALLOWED_RISK_LEVELS:
            raise ValueError(
                f"risk level {level!r} not in {cm.ALLOWED_RISK_LEVELS}"
            )
        ws[addr] = level

    # GL table (rows 38-40)
    for account_name, row_idx in cm.LEAD_GL_ROWS.items():
        gl = ctx.lead_gl_rows.get(account_name)
        if gl is None:
            continue
        for field_name, col in cm.LEAD_GL_COLS.items():
            value = getattr(gl, field_name)
            ws[f"{col}{row_idx}"] = value

    # Disclosure table (rows 63-66)
    for account_name, row_idx in cm.LEAD_DISCLOSURE_ROWS.items():
        current, prior = ctx.lead_disclosure.get(account_name, (0.0, 0.0))
        ws[f"{cm.LEAD_DISCLOSURE_CURRENT_COL}{row_idx}"] = current
        ws[f"{cm.LEAD_DISCLOSURE_PRIOR_COL}{row_idx}"] = prior

    # Restricted-cash items (rows 73+)
    for offset, item in enumerate(ctx.lead_restricted):
        row_idx = cm.LEAD_RESTRICTED_START_ROW + offset
        for field_name, col in cm.LEAD_RESTRICTED_COLS.items():
            ws[f"{col}{row_idx}"] = getattr(item, field_name)


def _write_recon_sheet(ws, ctx: FillContext) -> None:
    """Write C.02 Bank reconciliations."""
    if ctx.recon_account is None:
        logger.info("No reconciliation account in context; leaving C.02 blank")
        return

    info = ctx.recon_account
    ws[cm.RECON_ACCOUNT_SUBJECT] = info.subject
    ws[cm.RECON_BANK_NAME] = info.bank_name
    ws[cm.RECON_BANK_ACCOUNT] = info.bank_account
    ws[cm.RECON_CURRENCY] = info.currency
    ws[cm.RECON_STATEMENT_DATE] = info.statement_date

    ws[cm.RECON_BOOK_BASE] = ctx.recon_book_base
    ws[cm.RECON_BANK_BASE] = ctx.recon_bank_base

    # Four reconciliation buckets
    _write_recon_bucket(
        ws,
        items=ctx.recon_items_by_category.get("book_plus", []),
        row_range=cm.RECON_BOOK_PLUS_ROWS,
        desc_col=cm.RECON_BOOK_DESC_COL,
        amount_col=cm.RECON_BOOK_AMOUNT_COL,
        index_col=cm.RECON_BOOK_INDEX_COL,
    )
    _write_recon_bucket(
        ws,
        items=ctx.recon_items_by_category.get("book_minus", []),
        row_range=cm.RECON_BOOK_MINUS_ROWS,
        desc_col=cm.RECON_BOOK_DESC_COL,
        amount_col=cm.RECON_BOOK_AMOUNT_COL,
        index_col=cm.RECON_BOOK_INDEX_COL,
    )
    _write_recon_bucket(
        ws,
        items=ctx.recon_items_by_category.get("bank_plus", []),
        row_range=cm.RECON_BANK_PLUS_ROWS,
        desc_col=cm.RECON_BANK_DESC_COL,
        amount_col=cm.RECON_BANK_AMOUNT_COL,
        index_col=cm.RECON_BANK_INDEX_COL,
    )
    _write_recon_bucket(
        ws,
        items=ctx.recon_items_by_category.get("bank_minus", []),
        row_range=cm.RECON_BANK_MINUS_ROWS,
        desc_col=cm.RECON_BANK_DESC_COL,
        amount_col=cm.RECON_BANK_AMOUNT_COL,
        index_col=cm.RECON_BANK_INDEX_COL,
    )


def _write_recon_bucket(ws, items, row_range, desc_col, amount_col, index_col) -> None:
    """Write one of the four 4-row reconciliation buckets."""
    start, end = row_range
    capacity = end - start + 1
    for offset, item in enumerate(items[:capacity]):
        row_idx = start + offset
        ws[f"{desc_col}{row_idx}"] = item.description
        ws[f"{amount_col}{row_idx}"] = item.amount
        ws[f"{index_col}{row_idx}"] = item.index


def _write_cutoff_sheet(ws, ctx: FillContext) -> None:
    """Write C.03 Cutoff — pre/post-period sample tables + cutoff window."""
    ws[cm.CUTOFF_WINDOW] = ctx.cutoff_window
    _write_cutoff_table(ws, ctx.cutoff_pre_period, cm.CUTOFF_PRE_PERIOD_ROW_RANGE)
    _write_cutoff_table(ws, ctx.cutoff_post_period, cm.CUTOFF_POST_PERIOD_ROW_RANGE)


def _write_cutoff_table(ws, samples, row_range) -> None:
    start, end = row_range
    capacity = end - start + 1
    for offset, sample in enumerate(samples[:capacity]):
        row_idx = start + offset
        for field_name, col in cm.CUTOFF_SAMPLE_COLS.items():
            ws[f"{col}{row_idx}"] = getattr(sample, field_name)


def _write_clean_workbook(wb, ctx: FillContext) -> None:
    """Write the neutral clean C cash template."""
    _write_clean_lead_sheet(_clean_sheet(wb, "lead"), ctx)
    _write_clean_bkd_sheet(_clean_sheet(wb, "bkd"), ctx)
    _write_clean_recon_sheet(_clean_sheet(wb, "recon"), ctx)
    _write_clean_cutoff_sheet(_clean_sheet(wb, "cutoff"), ctx)


def _write_clean_lead_sheet(ws, ctx: FillContext) -> None:
    _set_cell(ws, cm_clean.LEAD_COMPANY_NAME, ctx.company_name)
    _set_cell(ws, cm_clean.LEAD_PERIOD_END, ctx.period_end)
    _set_cell(ws, cm_clean.LEAD_ANALYSIS_DATE, ctx.analysis_date)
    _set_cell(ws, cm_clean.LEAD_TE, ctx.te)
    _set_cell(ws, cm_clean.LEAD_SAD, ctx.sad)
    _set_cell(ws, cm_clean.LEAD_GAAP, ctx.gaap)
    _set_cell(ws, cm_clean.LEAD_AUDIT_STANDARD, "中国注册会计师审计准则")
    _set_cell(ws, cm_clean.LEAD_CURRENCY, ctx.currency)
    _set_cell(ws, cm_clean.LEAD_VARIATION_PCT, ctx.variation_pct)

    legacy_risk_cells = (
        cm.LEAD_RISK_COMPLETENESS,
        cm.LEAD_RISK_EXISTENCE,
        cm.LEAD_RISK_VALUATION,
        cm.LEAD_RISK_RIGHTS,
        cm.LEAD_RISK_PRESENTATION,
    )
    for legacy_addr, clean_addr in zip(legacy_risk_cells, cm_clean.LEAD_RISK_CELLS):
        level = ctx.risk_levels.get(legacy_addr, "不适用")
        _set_cell(ws, clean_addr, _translate_clean_risk(level))

    for account_name, row_idx in cm_clean.LEAD_GL_ROWS.items():
        gl = ctx.lead_gl_rows.get(account_name)
        if gl is None:
            continue
        values = {
            "bookkeeping_code": gl.bookkeeping_code or ctx.company_name,
            "gl_account_code": gl.gl_account_code,
            "account_name": account_name,
            "index_ref": gl.index_ref,
            "book_value_unaudited": gl.book_value_unaudited,
            "book_adjustment": gl.book_adjustment,
            "audit_adjustment": gl.audit_adjustment,
            "prior_year_audited": gl.prior_year_audited,
            "notes_tag": ctx.lead_gl_notes.get(account_name, gl.notes_tag),
        }
        for field_name, col in cm_clean.LEAD_GL_COLS.items():
            _set_cell(ws, f"{col}{row_idx}", values[field_name])

    restricted_total = sum(item.amount for item in ctx.lead_restricted)
    if restricted_total:
        notes = "; ".join(
            f"{item.description} {item.amount:,.2f} {item.nature}".strip()
            for item in ctx.lead_restricted
        )
        _set_cell(ws, cm_clean.LEAD_RESTRICTED_CURRENT, restricted_total)
        _set_cell(ws, cm_clean.LEAD_RESTRICTED_PRIOR, 0.0)
        _set_cell(ws, cm_clean.LEAD_RESTRICTED_NOTE, notes)


def _write_clean_bkd_sheet(ws, ctx: FillContext) -> None:
    for offset, row in enumerate(ctx.period_summary_rows[: cm_clean.BKD_MAX_ROWS]):
        row_idx = cm_clean.BKD_START_ROW + offset
        values = {
            "company_name": ctx.company_name,
            "account_name": row.account_name,
            "bank_name": row.bank_name or "现金",
            "bank_account": row.bank_account or "N/A",
            "currency": row.currency,
            "period_end_balance_fx": row.period_end_balance_fx,
            "period_end_balance_local": row.period_end_balance_local,
            "account_usage": _account_usage(row),
            "confirmation_required": _confirmation_required(row),
            "alternative_procedure": _alternative_procedure(row),
            "notes": row.restriction_note if row.is_restricted else "",
        }
        for field_name, col in cm_clean.BKD_COLS.items():
            _set_cell(ws, f"{col}{row_idx}", values[field_name])


def _write_clean_recon_sheet(ws, ctx: FillContext) -> None:
    if ctx.recon_account is None:
        return

    info = ctx.recon_account
    _set_cell(ws, cm_clean.RECON_ACCOUNT_SUBJECT, info.subject)
    _set_cell(ws, cm_clean.RECON_INDEX_REF, "C.00 BKD")
    _set_cell(ws, cm_clean.RECON_STATEMENT_DATE, info.statement_date)
    _set_cell(ws, cm_clean.RECON_BANK_NAME, info.bank_name)
    _set_cell(ws, cm_clean.RECON_BANK_ACCOUNT, info.bank_account)
    _set_cell(ws, cm_clean.RECON_CURRENCY, info.currency)
    _set_cell(ws, cm_clean.RECON_ACCOUNT_NATURE, "经营收付款账户")
    _set_cell(
        ws,
        cm_clean.RECON_HAS_ITEMS,
        "是" if any(ctx.recon_items_by_category.values()) else "否",
    )
    _set_cell(ws, cm_clean.RECON_SOURCE, "PBC / 函证或银行回函 / 银行流水")
    if ctx.recon_rationale:
        _set_cell(ws, cm_clean.RECON_SELECTION_RATIONALE, ctx.recon_rationale)
    if ctx.recon_conclusion:
        _set_cell(ws, cm_clean.RECON_CONCLUSION, ctx.recon_conclusion)
    _set_cell(ws, cm_clean.RECON_BOOK_BASE, ctx.recon_book_base)
    _set_cell(ws, cm_clean.RECON_BANK_BASE, ctx.recon_bank_base)

    _write_clean_recon_bucket(
        ws,
        items=ctx.recon_items_by_category.get("book_plus", []),
        row_range=cm_clean.RECON_BOOK_PLUS_ROWS,
        desc_col=cm_clean.RECON_BOOK_DESC_COL,
        amount_col=cm_clean.RECON_BOOK_AMOUNT_COL,
        index_col=cm_clean.RECON_BOOK_INDEX_COL,
    )
    _write_clean_recon_bucket(
        ws,
        items=ctx.recon_items_by_category.get("book_minus", []),
        row_range=cm_clean.RECON_BOOK_MINUS_ROWS,
        desc_col=cm_clean.RECON_BOOK_DESC_COL,
        amount_col=cm_clean.RECON_BOOK_AMOUNT_COL,
        index_col=cm_clean.RECON_BOOK_INDEX_COL,
    )
    _write_clean_recon_bucket(
        ws,
        items=ctx.recon_items_by_category.get("bank_plus", []),
        row_range=cm_clean.RECON_BANK_PLUS_ROWS,
        desc_col=cm_clean.RECON_BANK_DESC_COL,
        amount_col=cm_clean.RECON_BANK_AMOUNT_COL,
        index_col=cm_clean.RECON_BANK_INDEX_COL,
    )
    _write_clean_recon_bucket(
        ws,
        items=ctx.recon_items_by_category.get("bank_minus", []),
        row_range=cm_clean.RECON_BANK_MINUS_ROWS,
        desc_col=cm_clean.RECON_BANK_DESC_COL,
        amount_col=cm_clean.RECON_BANK_AMOUNT_COL,
        index_col=cm_clean.RECON_BANK_INDEX_COL,
    )


def _write_clean_recon_bucket(ws, items, row_range, desc_col, amount_col, index_col) -> None:
    start, end = row_range
    capacity = end - start + 1
    for offset, item in enumerate(items[:capacity]):
        row_idx = start + offset
        _set_cell(ws, f"{desc_col}{row_idx}", item.description)
        _set_cell(ws, f"{amount_col}{row_idx}", item.amount)
        _set_cell(ws, f"{index_col}{row_idx}", item.index)


def _write_clean_cutoff_sheet(ws, ctx: FillContext) -> None:
    _set_cell(ws, cm_clean.CUTOFF_WINDOW, ctx.cutoff_window)
    _set_cell(ws, cm_clean.CUTOFF_WORKDAYS, 5)
    _set_cell(
        ws,
        cm_clean.CUTOFF_REASON,
        ctx.cutoff_reason or "根据期末前后银行间转账风险及交易处理周期确定测试期间。",
    )
    _set_cell(
        ws,
        cm_clean.CUTOFF_CONCLUSION,
        ctx.cutoff_conclusion or "已根据样本执行截止性测试，关注跨期转账是否恰当反映于余额调节表。",
    )
    account_bank = {
        row.bank_account: row.bank_name
        for row in ctx.period_summary_rows
        if row.bank_account and row.bank_name
    }
    _write_clean_cutoff_table(ws, ctx.cutoff_pre_period, cm_clean.CUTOFF_PRE_PERIOD_ROW_RANGE, account_bank)
    _write_clean_cutoff_table(ws, ctx.cutoff_post_period, cm_clean.CUTOFF_POST_PERIOD_ROW_RANGE, account_bank)


def _write_clean_cutoff_table(ws, samples, row_range, account_bank: dict[str, str]) -> None:
    start, end = row_range
    capacity = end - start + 1
    for offset, sample in enumerate(samples[:capacity]):
        row_idx = start + offset
        values = _clean_cutoff_values(sample, account_bank)
        for field_name, col in cm_clean.CUTOFF_SAMPLE_COLS.items():
            _set_cell(ws, f"{col}{row_idx}", values[field_name])


def _clean_cutoff_values(sample, account_bank: dict[str, str]) -> dict[str, object]:
    values = {field_name: getattr(sample, field_name, "") for field_name in cm.CUTOFF_SAMPLE_COLS}
    values["accounting_ok"] = "是"

    if sample.out_amount and not sample.in_amount:
        counterparty = sample.in_bank_name
        if counterparty in account_bank:
            values["in_bank_account"] = counterparty
            values["in_bank_name"] = account_bank[counterparty]
        values["in_amount"] = abs(sample.out_amount)
    elif sample.in_amount and not sample.out_amount:
        counterparty = sample.in_bank_account
        if counterparty in account_bank:
            values["out_bank_account"] = counterparty
            values["out_bank_name"] = account_bank[counterparty]
        values["out_amount"] = -abs(sample.in_amount)

    return values


def _translate_clean_risk(level: str) -> str:
    translated = cm_clean.RISK_LEVEL_TRANSLATION.get(level)
    if translated is None:
        raise ValueError(f"risk level {level!r} not in clean template map")
    return translated


def _account_usage(row: PeriodSummaryRow) -> str:
    if row.account_name == "库存现金":
        return "库存现金及零星备用金"
    if row.is_restricted:
        return row.restriction_note or "受限货币资金"
    if row.account_name == "其他货币资金":
        return "保证金、平台备付金或其他货币资金"
    return "经营收付款账户"


def _confirmation_required(row: PeriodSummaryRow) -> str:
    if row.account_name == "库存现金" or not row.bank_account:
        return "不适用"
    return "是"


def _alternative_procedure(row: PeriodSummaryRow) -> str:
    if _confirmation_required(row) == "是":
        return ""
    return "执行现金盘点或替代核对程序"


# ──────────────────────────────────────────────────────────────────────────────
# Top-level API
# ──────────────────────────────────────────────────────────────────────────────


def fill_cash_workpaper(
    materials_dir: Path | str,
    template_path: Path | str,
    output_path: Path | str,
    *,
    llm_enhance: bool = False,
) -> Path:
    """Convenience API: load materials → build context → write filled xlsx.

    Returns the output Path. Raises on any I/O or schema error.
    """
    materials_dir = Path(materials_dir)
    template_path = Path(template_path)
    output_path = Path(output_path)

    if not template_path.exists():
        raise FileNotFoundError(f"template not found: {template_path}")

    logger.info("Loading materials from %s", materials_dir)
    materials = load_case_materials(materials_dir)
    logger.info(
        "Loaded %d bank-statement rows, %d GL rows, %d confirmations, "
        "%d period-summary rows, %d reconciliation items",
        len(materials.bank_statements),
        len(materials.gl_bank),
        len(materials.confirmations),
        len(materials.period_summary),
        len(materials.reconciliation_items),
    )

    context = build_fill_context(materials)
    if llm_enhance:
        from .llm_enhancer import enhance_fill_context_with_llm

        context = enhance_fill_context_with_llm(context)

    logger.info("Writing filled workpaper to %s", output_path)
    return write_cells_to_workbook(template_path, output_path, context)


__all__ = [
    "FIXED_GL_ACCOUNTS",
    "DEFAULT_RISK_LEVELS",
    "DEFAULT_CUTOFF_WINDOW",
    "ReconAccountInfo",
    "FillContext",
    "build_fill_context",
    "write_cells_to_workbook",
    "fill_cash_workpaper",
]
