# -*- coding: utf-8 -*-
from __future__ import annotations

import argparse
import json
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from audit_rag.agents import AgentTurn
from audit_rag.config import get_settings
from audit_rag.data_tools import Finding, Voucher
from audit_rag.pipeline import DEFAULT_CASE, run_audit_pipeline
from audit_rag.rag import RetrievedChunk, format_chunk_citation, format_rag_context


DEFAULT_TEMPLATE_KEYWORD = "K1 SWP 固定资产"
AI_SHEET_NAME = "AI_Audit_Assistant"
DEFAULT_GAAP = "企业会计准则"
DEFAULT_CURRENCY = "人民币"

# Per-case-type defaults. The cash workflow defaults to the neutral optimized
# five-sheet template and reads client/period/currency from case_metadata.json
# inside the materials dir, so it intentionally ignores --client-name / --gaap
# / --currency.
CASE_TYPE_DEFAULTS = {
    "fixed_asset": {
        "template_keyword": "K1 SWP 固定资产",
    },
    "cash": {
        "template_keyword": "核心优化版",
    },
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate an Excel workpaper copy from a standard SWP template.")
    parser.add_argument(
        "--case-type",
        choices=sorted(CASE_TYPE_DEFAULTS.keys()),
        default="fixed_asset",
        help="Which workpaper to fill. 'fixed_asset' uses the agent pipeline (legacy);"
        " 'cash' uses the benchmark materials directory (new in Phase 1).",
    )
    parser.add_argument(
        "--materials-dir",
        default="",
        help="For --case-type=cash: path to a benchmarks/materials/case_xxx/ folder."
        " Ignored for fixed_asset.",
    )
    parser.add_argument("--mode", choices=["mock", "autogen"], default="mock", help="mock is free; autogen calls API.")
    parser.add_argument("--case", default=DEFAULT_CASE, help="Audit task description (fixed_asset only).")
    parser.add_argument("--voucher-file", default="", help="Optional voucher CSV path (fixed_asset only).")
    parser.add_argument("--template-root", required=True, help="Folder that contains standard SWP templates.")
    parser.add_argument(
        "--template-keyword",
        default="",
        help="Override the template filename keyword. Defaults to the keyword for the selected --case-type.",
    )
    parser.add_argument("--client-name", default="XYZ公司", help="Client name used in output filename (fixed_asset only).")
    parser.add_argument("--gaap", default=DEFAULT_GAAP, help="GAAP label (fixed_asset only).")
    parser.add_argument("--currency", default=DEFAULT_CURRENCY, help="Currency label (fixed_asset only).")
    args = parser.parse_args()

    # Fill in the per-case-type default keyword if user didn't override.
    if not args.template_keyword:
        args.template_keyword = CASE_TYPE_DEFAULTS[args.case_type]["template_keyword"]

    # Validate cash-specific args
    if args.case_type == "cash" and not args.materials_dir:
        parser.error("--case-type=cash requires --materials-dir")

    return args


def find_template(template_root: Path, keyword: str) -> Path:
    if not template_root.exists():
        raise FileNotFoundError(template_root)

    candidates = [
        path
        for path in template_root.rglob("*")
        if path.is_file() and path.suffix.lower() in {".xlsx", ".xlsm"} and keyword in path.name
    ]
    if not candidates:
        raise FileNotFoundError(f"No Excel template filename contains: {keyword}")
    return sorted(candidates, key=lambda p: (len(p.name), p.name))[0]


def build_agent_outputs(
    mode: str,
    case_description: str,
    voucher_file: str,
) -> tuple[list[Voucher], list[Finding], list[RetrievedChunk], list[AgentTurn], Path]:
    result = run_audit_pipeline(mode=mode, case_description=case_description, voucher_file=voucher_file)
    return result.vouchers, result.findings, result.rag_chunks, result.turns, result.report_path


def copy_template(template_path: Path, client_name: str, mode: str) -> Path:
    settings = get_settings()
    workpapers_dir = settings.project_root / "output" / "workpapers"
    workpapers_dir.mkdir(parents=True, exist_ok=True)

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")[:-3]
    safe_client = client_name.replace("/", "_").replace("\\", "_").strip() or "Client"
    suffix = template_path.suffix
    output_name = f"{stamp}_{template_path.stem}_{safe_client}_{mode}_AI底稿{suffix}"
    output_path = workpapers_dir / output_name
    shutil.copy2(template_path, output_path)
    return output_path


def write_ai_sheet(
    workbook_path: Path,
    case_description: str,
    mode: str,
    client_name: str,
    gaap: str,
    currency: str,
    vouchers: list[Voucher],
    findings: list[Finding],
    rag_chunks: list[RetrievedChunk],
    turns: list[AgentTurn],
    report_path: Path,
) -> None:
    keep_vba = workbook_path.suffix.lower() == ".xlsm"
    wb = load_workbook(workbook_path, keep_vba=keep_vba)
    template_map = load_template_map()
    write_fixed_asset_template_fields(
        wb=wb,
        template_map=template_map["fixed_asset"],
        case_description=case_description,
        client_name=client_name,
        gaap=gaap,
        currency=currency,
        vouchers=vouchers,
        findings=findings,
        turns=turns,
        rag_chunks=rag_chunks,
        report_path=report_path,
    )

    if AI_SHEET_NAME in wb.sheetnames:
        del wb[AI_SHEET_NAME]
    ws = wb.create_sheet(AI_SHEET_NAME, 0)

    title_fill = PatternFill("solid", fgColor="1F4E79")
    title_font = Font(color="FFFFFF", bold=True, size=14)
    header_font = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    ws["A1"] = "AI 审计助手工作底稿摘要"
    ws["A1"].fill = title_fill
    ws["A1"].font = title_font
    ws.merge_cells("A1:D1")

    metadata = [
        ("生成时间", datetime.now().isoformat(timespec="seconds")),
        ("运行模式", mode),
        ("样本凭证数", len(vouchers)),
        ("异常发现数", len(findings)),
        ("Markdown 底稿", str(report_path)),
        ("审计任务", case_description),
    ]
    row = 3
    for key, value in metadata:
        ws.cell(row=row, column=1, value=key).font = header_font
        ws.cell(row=row, column=2, value=value).alignment = wrap
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        row += 1

    row += 1
    row = write_section(ws, row, "数据提取发现")
    for idx, finding in enumerate(findings, 1):
        values = [
            idx,
            finding.risk_level,
            finding.issue,
            f"{finding.evidence}\n建议程序: {finding.suggested_procedure}",
        ]
        for col, value in enumerate(values, 1):
            ws.cell(row=row, column=col, value=value).alignment = wrap
        row += 1

    row += 1
    row = write_section(ws, row, "RAG 检索依据")
    ws.cell(row=row, column=1, value=format_rag_context(rag_chunks)).alignment = wrap
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 2

    for turn in turns:
        row = write_section(ws, row, turn.speaker)
        ws.cell(row=row, column=1, value=turn.content).alignment = wrap
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 2

    for col, width in {"A": 18, "B": 18, "C": 28, "D": 90}.items():
        ws.column_dimensions[col].width = width
    for sheet_row in range(1, row + 1):
        ws.row_dimensions[sheet_row].height = 36
    ws.freeze_panes = "A3"

    wb.save(workbook_path)


def load_template_map() -> dict:
    settings = get_settings()
    map_path = settings.project_root / "data" / "template_field_map.json"
    return json.loads(map_path.read_text(encoding="utf-8"))


def write_fixed_asset_template_fields(
    wb,
    template_map: dict,
    case_description: str,
    client_name: str,
    gaap: str,
    currency: str,
    vouchers: list[Voucher],
    findings: list[Finding],
    turns: list[AgentTurn],
    rag_chunks: list[RetrievedChunk],
    report_path: Path,
) -> None:
    lead_map = template_map["lead_sheet"]
    additions_map = template_map["additions_test"]

    lead_ws = wb[lead_map["sheet"]]
    additions_ws = next(ws for ws in wb.worksheets if ws.title.startswith(additions_map["sheet_prefix"]))

    high_risk_findings = [finding for finding in findings if finding.risk_level == "高"]
    selected_findings = high_risk_findings or findings[:1]
    selected_vouchers = match_vouchers(vouchers, selected_findings)
    capital_related_vouchers = find_capital_related_vouchers(vouchers)

    lead_ws[lead_map["client_name"]] = client_name
    lead_ws[lead_map["analysis_date"]] = datetime.now().date()
    lead_ws[lead_map["gaap"]] = gaap
    lead_ws[lead_map["currency"]] = currency
    lead_ws[lead_map["ai_fluctuation_note"]] = build_lead_sheet_note(
        case_description=case_description,
        vouchers=vouchers,
        findings=findings,
        selected_findings=selected_findings,
        rag_chunks=rag_chunks,
        report_path=report_path,
    )
    lead_ws[lead_map["ai_fluctuation_note"]].alignment = Alignment(wrap_text=True, vertical="top")

    additions_ws[additions_map["population_description"]] = build_population_description(vouchers, capital_related_vouchers)
    additions_ws[additions_map["key_item_rationale"]] = build_key_item_rationale(selected_findings, rag_chunks)
    additions_ws[additions_map["purchase_total"]] = sum(v.amount for v in capital_related_vouchers)
    additions_ws[additions_map["key_items_tested"]] = sum(v.amount for v in selected_vouchers)

    for cell_ref in [additions_map["population_description"], additions_map["key_item_rationale"]]:
        additions_ws[cell_ref].alignment = Alignment(wrap_text=True, vertical="top")

    first_row = int(additions_map["first_sample_row"])
    max_rows = int(additions_map["max_sample_rows"])
    for offset, voucher in enumerate(selected_vouchers[:max_rows]):
        finding = next((item for item in selected_findings if item.voucher_id == voucher.voucher_id), None)
        write_addition_sample_row(additions_ws, first_row + offset, voucher, finding, rag_chunks)


def match_vouchers(vouchers: list[Voucher], findings: list[Finding]) -> list[Voucher]:
    ids = {finding.voucher_id for finding in findings}
    return [voucher for voucher in vouchers if voucher.voucher_id in ids]


def find_capital_related_vouchers(vouchers: list[Voucher]) -> list[Voucher]:
    capital_keywords = ["设备", "服务器", "机器", "固定资产", "工程", "装修", "硬件", "车辆", "产线"]
    return [
        voucher
        for voucher in vouchers
        if any(
            keyword in f"{voucher.debit_subject} {voucher.summary} {voucher.attachment}"
            for keyword in capital_keywords
        )
    ]


def build_lead_sheet_note(
    case_description: str,
    vouchers: list[Voucher],
    findings: list[Finding],
    selected_findings: list[Finding],
    rag_chunks: list[RetrievedChunk],
    report_path: Path,
) -> str:
    high_risk_total = sum(
        voucher.amount
        for voucher in vouchers
        if any(finding.voucher_id == voucher.voucher_id and finding.risk_level == "高" for finding in selected_findings)
    )
    finding_lines = "\n".join(
        f"- {finding.voucher_id}: [{finding.risk_level}风险] {finding.issue}。{finding.suggested_procedure}"
        for finding in selected_findings
    )
    return (
        "AI 辅助说明: 基于样例凭证和当前 RAG 片段执行固定资产相关风险扫描。\n"
        f"审计任务: {case_description}\n"
        f"样本凭证数: {len(vouchers)}；识别发现数: {len(findings)}；高风险样本金额: {high_risk_total:,.2f}。\n"
        f"{finding_lines}\n"
        "初步结论: 发现疑似资本性支出费用化事项，可能导致固定资产低估、管理费用高估，并影响折旧和期间损益。"
        f"\n参考依据: {build_rag_reference_brief(rag_chunks)}"
        f"\n详细 AI 输出见: {report_path}"
    )


def build_population_description(vouchers: list[Voucher], capital_related_vouchers: list[Voucher]) -> str:
    total = sum(v.amount for v in capital_related_vouchers)
    ids = ", ".join(v.voucher_id for v in capital_related_vouchers) or "无"
    return (
        "AI 辅助填列: 本次以样例凭证作为新增固定资产测试的初始总体，筛选摘要、科目或附件中包含"
        "“设备、服务器、固定资产、硬件”等资本化关键词的项目。"
        f"\n总体凭证数: {len(vouchers)}；资本相关凭证数: {len(capital_related_vouchers)}；金额合计: {total:,.2f}。"
        f"\n纳入总体的凭证: {ids}。"
    )


def build_key_item_rationale(findings: list[Finding], rag_chunks: list[RetrievedChunk]) -> str:
    if not findings:
        return "AI 辅助填列: 当前样本未识别需作为关键项目单独测试的高风险固定资产新增事项。"

    lines = [
        "AI 辅助填列: 选择以下项目作为关键项目，原因是金额重大且摘要/附件显示可能形成长期资产，但账务处理存在费用化或分类异常风险。"
    ]
    for finding in findings:
        lines.append(f"- {finding.voucher_id}: [{finding.risk_level}风险] {finding.issue}；证据: {finding.evidence}")
    lines.append(f"参考依据: {build_rag_reference_brief(rag_chunks)}")
    return "\n".join(lines)


def write_addition_sample_row(ws, row: int, voucher: Voucher, finding: Finding | None, rag_chunks: list[RetrievedChunk]) -> None:
    issue = finding.issue if finding else "资本相关样本"
    evidence = finding.evidence if finding else f"{voucher.summary}; {voucher.attachment}"
    suggested = finding.suggested_procedure if finding else "检查合同、发票、验收和资产台账。"
    values = {
        "B": "关键项目" if finding and finding.risk_level == "高" else "代表性样本",
        "C": "电子设备/服务器",
        "D": voucher.debit_subject,
        "E": voucher.voucher_id,
        "F": voucher.summary,
        "G": voucher.amount,
        "H": voucher.date,
        "I": "待从资产台账确认",
        "J": "待确认",
        "K": "待确认",
        "L": voucher.amount,
        "M": (
            f"{evidence}\n取得/拟检查证据: {voucher.attachment}\n建议程序: {suggested}"
            f"\n参考依据: {build_rag_reference_brief(rag_chunks)}"
        ),
        "O": "否 - 当前账务存在费用化/分类异常" if finding else "待确认",
        "P": "待与合同/发票核对",
        "Q": "待确认使用寿命和折旧政策",
        "R": f"异常: {issue}" if finding else "待确认分类",
    }
    for col, value in values.items():
        cell = ws[f"{col}{row}"]
        cell.value = value
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def build_rag_reference_brief(rag_chunks: list[RetrievedChunk], limit: int = 2) -> str:
    if not rag_chunks:
        return "当前未检索到可引用依据。"
    selected = rag_chunks[:limit]
    return "；".join(format_chunk_citation(chunk) for chunk in selected)


def write_section(ws, row: int, title: str) -> int:
    ws.cell(row=row, column=1, value=title)
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="D9EAF7")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    return row + 1


def main() -> None:
    args = parse_args()
    if args.case_type == "cash":
        _run_cash_workflow(args)
    else:
        _run_fixed_asset_workflow(args)


def _run_fixed_asset_workflow(args: argparse.Namespace) -> None:
    """Legacy workflow: runs the audit pipeline then writes the K1 固定资产 template."""
    template_path = find_template(Path(args.template_root), args.template_keyword)
    vouchers, findings, rag_chunks, turns, report_path = build_agent_outputs(
        args.mode, args.case, args.voucher_file
    )
    workbook_path = copy_template(template_path, args.client_name, args.mode)
    write_ai_sheet(
        workbook_path,
        args.case,
        args.mode,
        args.client_name,
        args.gaap,
        args.currency,
        vouchers,
        findings,
        rag_chunks,
        turns,
        report_path,
    )

    print("Workpaper generated.")
    print(f"Mode: {args.mode}")
    print(f"Template: {template_path}")
    print(f"Markdown report: {report_path}")
    print(f"Excel workpaper: {workbook_path}")


def _run_cash_workflow(args: argparse.Namespace) -> None:
    """Phase 1 cash workflow.

    Reads client materials from ``args.materials_dir``, runs the rule-based
    cash workpaper filler from ``benchmarks.agent.cash_workpaper_filler``,
    and writes the result alongside other workpapers in ``output/workpapers/``.

    The audit pipeline (DeepSeek / mock agents) is intentionally NOT invoked:
    the cash workflow's source of truth is the materials directory plus the
    cell map, not LLM output. Phase 3 layers LLM reasoning on top.
    """
    # Lazy import so fixed_asset users aren't forced to load openpyxl-heavy
    # benchmark modules.
    from benchmarks.agent.cash_workpaper_filler import fill_cash_workpaper
    from benchmarks.agent.materials_loader import load_case_materials

    materials_dir = Path(args.materials_dir)
    if not materials_dir.is_dir():
        raise NotADirectoryError(f"--materials-dir does not exist: {materials_dir}")

    # Sanity-load metadata so we fail fast on schema errors (and to get a
    # nicer client_name for the output filename).
    materials = load_case_materials(materials_dir)
    client_name = materials.meta.client_name

    template_path = find_template(Path(args.template_root), args.template_keyword)
    output_path = _build_cash_output_path(template_path, client_name, materials.meta.case_id)

    llm_enhance = args.mode == "autogen"
    fill_cash_workpaper(materials_dir, template_path, output_path, llm_enhance=llm_enhance)

    print("Cash workpaper generated.")
    print(f"Case ID: {materials.meta.case_id}")
    print(f"Client: {client_name}")
    print(f"Template: {template_path}")
    print(f"Materials: {materials_dir}")
    print(f"LLM: {'enabled' if llm_enhance else 'disabled'}")
    print(f"Excel workpaper: {output_path}")


def _build_cash_output_path(template_path: Path, client_name: str, case_id: str) -> Path:
    settings = get_settings()
    workpapers_dir = settings.project_root / "output" / "workpapers"
    workpapers_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")[:-3]
    safe_client = client_name.replace("/", "_").replace("\\", "_").strip() or "Client"
    safe_case = case_id.replace("/", "_").replace("\\", "_").strip() or "case"
    suffix = template_path.suffix
    return workpapers_dir / f"{stamp}_{template_path.stem}_{safe_client}_{safe_case}_AI底稿{suffix}"


if __name__ == "__main__":
    main()
